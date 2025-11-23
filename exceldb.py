import pandas as pd
import random
from datetime import datetime, timedelta
from openpyxl import Workbook

# --- Données de base ---
commerciaux = ["Moussa", "Mariam", "Adama", "Fatou", "Issa", "Van", "Oumou", "Mamy"]
communes = ["Kalaban", "Lafiabougou", "ACI 2000", "Sogoniko", "Missabougou", "Banconi", "Baco-Djicoroni", "Hamdallaye"]
contacts = ["70352450", "65214789", "76459812", "74125698", "76841235", "70984521", "72561489", "78451236"]

# --- Période ---
start_date = datetime(2025, 9, 1)
end_date = datetime(2025, 11, 30)

# --- Génération de données fictives ---
data = []
for _ in range(500):
    date = start_date + timedelta(days=random.randint(0, (end_date - start_date).days))
    commercial = random.choice(commerciaux)
    contact = random.choice(contacts)
    quantite = random.randint(5, 50)
    prix_unitaire = random.choice([200, 250, 300, 350, 400])
    montant_total = quantite * prix_unitaire
    commune = random.choice(communes)
    commentaire = random.choice(["", "Client fidèle", "Promo", "Commande spéciale", "Paiement différé"])
    data.append([date.strftime("%d/%m/%Y"), commercial, contact, quantite, prix_unitaire, montant_total, commune, commentaire])

# --- Création du DataFrame ---
df = pd.DataFrame(data, columns=[
    "Date", "Commercial", "Contact", "Quantité", "Prix unitaire", 
    "Montant total", "Commune", "Commentaire"
])

# --- Création du fichier Excel ---
file_path = "Ventes_Stats_Complet.xlsx"

with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Ventes", index=False)

    # Créer une feuille Stats
    wb = writer.book
    ws = wb.create_sheet("Stats")

    # En-têtes
    ws["A1"] = "📊 Statistiques Globales"
    ws["A3"] = "Total des ventes (FCFA)"
    ws["B3"] = "=SOMME(Ventes!F2:F501)"
    ws["A4"] = "Nombre de ventes"
    ws["B4"] = "=NB(Ventes!F2:F501)"
    ws["A5"] = "Quantité totale vendue"
    ws["B5"] = "=SOMME(Ventes!D2:D501)"
    ws["A6"] = "Prix unitaire moyen"
    ws["B6"] = "=MOYENNE(Ventes!E2:E501)"

    ws["A8"] = "🏆 Meilleur commercial du mois"
    ws["B8"] = "=INDEX(Ventes!B2:B501;EQUIV(MAX(SOMME.SI(Ventes!B2:B501;Ventes!B2:B501;Ventes!F2:F501));SOMME.SI(Ventes!B2:B501;Ventes!B2:B501;Ventes!F2:F501);0))"

    ws["A10"] = "📅 Ventes par mois"
    ws["A11"] = "Septembre 2025"
    ws["B11"] = "=SOMMEPROD((MOIS(DATEVAL(Ventes!A2:A501))=9)*(Ventes!F2:F501))"
    ws["A12"] = "Octobre 2025"
    ws["B12"] = "=SOMMEPROD((MOIS(DATEVAL(Ventes!A2:A501))=10)*(Ventes!F2:F501))"
    ws["A13"] = "Novembre 2025"
    ws["B13"] = "=SOMMEPROD((MOIS(DATEVAL(Ventes!A2:A501))=11)*(Ventes!F2:F501))"

print(f"✅ Fichier généré avec succès : {file_path}")
