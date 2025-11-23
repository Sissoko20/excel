from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import LineChart, BarChart, Reference

OUTPUT_FILE = "Modele_Station_Dashboard.xlsx"
NB_LIGNES = 200

wb = Workbook()
ws = wb.active
ws.title = "Gestion Carburant"

# -------------------------
# Paramètres en haut
# -------------------------
ws["A1"] = "Paramètres"
ws["A1"].font = Font(bold=True, size=14)

ws["A2"] = "Stock initial (L)"
ws["B2"] = 9000
ws["A3"] = "PU (Prix Unitaire)"
ws["B3"] = 775
ws["A4"] = "Seuil sécurité (L)"
ws["B4"] = 1000

# -------------------------
# Colonnes principales
# -------------------------
headers = ["Date","N° Engin","PU","Montant","Quantité achetée (L)",
           "Stock initial (L)","Stock réel (L)","Alerte"]

for col_idx, h in enumerate(headers, start=1):
    cell = ws.cell(row=6, column=col_idx, value=h)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

ws.freeze_panes = "A7"

# Formules ligne 7
ws["C7"] = "=B3"                         # PU paramétré
ws["E7"] = "=IF(C7>0,D7/C7,0)"           # Quantité achetée
ws["D7"] = "=C7*E7"                      # Montant
ws["F7"] = "=B2"                         # Stock initial paramétré
ws["G7"] = "=F7-E7"                      # Stock réel
ws["H7"] = '=IF(G7<B4,"ALERTE","OK")'

# Recopie formules pour les lignes suivantes
for r in range(8, NB_LIGNES+7):
    ws[f"C{r}"] = "=B3"
    ws[f"E{r}"] = f"=IF(C{r}>0,D{r}/C{r},0)"
    ws[f"D{r}"] = f"=C{r}*E{r}"
    ws[f"F{r}"] = "=B2"
    ws[f"G{r}"] = f"=F{r}-E{r}"
    ws[f"H{r}"] = f'=IF(G{r}<B4,"ALERTE","OK")'

# Mise en forme conditionnelle
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

rule_alert = FormulaRule(formula=['$H7="ALERTE"'], fill=red_fill)
rule_ok = FormulaRule(formula=['$H7="OK"'], fill=green_fill)

ws.conditional_formatting.add(f"H7:H{NB_LIGNES+6}", rule_alert)
ws.conditional_formatting.add(f"H7:H{NB_LIGNES+6}", rule_ok)

# Graphique 1 : Stock réel dans le temps
line_chart = LineChart()
line_chart.title = "Évolution du Stock réel"
line_chart.y_axis.title = "Litres"
line_chart.x_axis.title = "Date"
data_ref = Reference(ws, min_col=7, min_row=6, max_row=NB_LIGNES+6)  # Stock réel
dates_ref = Reference(ws, min_col=1, min_row=7, max_row=NB_LIGNES+6) # Dates
line_chart.add_data(data_ref, titles_from_data=True)
line_chart.set_categories(dates_ref)
ws.add_chart(line_chart, "J2")

# Graphique 2 : Quantité achetée par engin
bar_chart = BarChart()
bar_chart.title = "Quantité achetée par Engin"
bar_chart.y_axis.title = "Litres"
bar_chart.x_axis.title = "N° Engin"
data_ref2 = Reference(ws, min_col=5, min_row=6, max_row=NB_LIGNES+6)  # Quantité achetée
engins_ref = Reference(ws, min_col=2, min_row=7, max_row=NB_LIGNES+6) # N° Engin
bar_chart.add_data(data_ref2, titles_from_data=True)
bar_chart.set_categories(engins_ref)
ws.add_chart(bar_chart, "J20")

# Largeur colonnes
widths = [14,14,10,14,18,18,18,12]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64+i)].width = w

# Sauvegarde
wb.save(OUTPUT_FILE)
print(f"Fichier généré: {OUTPUT_FILE}")
