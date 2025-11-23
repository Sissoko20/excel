from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import LineChart, Reference

OUTPUT_FILE = "Modele_Station_Carburant.xlsx"
NB_LIGNES_COMMANDE = 200
NB_LIGNES_VENTES = 500

wb = Workbook()

# -------------------------
# Feuille Paramètres
# -------------------------
ws_param = wb.active
ws_param.title = "Paramètres"

ws_param["A3"] = "Capacité cuve (L)"
ws_param["B3"] = 18000
ws_param["A4"] = "Niveau sécurité (%)"
ws_param["B4"] = 0.10
ws_param["A5"] = "Stock initial (L)"
ws_param["B5"] = 12000
ws_param["A6"] = "Seuil sécurité (L)"
ws_param["B6"] = "=B3*B4"

dv_capacite = DataValidation(type="list", formula1='"9000,18000,27000,45000"', allow_blank=False)
ws_param.add_data_validation(dv_capacite)
dv_capacite.add(ws_param["B3"])

# -------------------------
# Feuille Commandes
# -------------------------
ws_cmd = wb.create_sheet("Commandes")

headers = ["Date","N° Châssis","Montant","PU","Quantité livrée (L)",
           "Stock initial (L)","Quantité vendue (L)","Stock final (L)","Alerte"]

for col_idx, h in enumerate(headers, start=1):
    cell = ws_cmd.cell(row=1, column=col_idx, value=h)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

ws_cmd.freeze_panes = "A2"

# Formules ligne 2
ws_cmd["E2"] = "=IFERROR(C2/D2,0)"             # Quantité livrée
ws_cmd["F2"] = "=Paramètres!B5"                # Stock initial
ws_cmd["G2"] = '=IFERROR(SUMIF(Ventes!A:A,A2,Ventes!B:B),0)'  # Quantité vendue auto
ws_cmd["H2"] = "=MAX(0,F2+E2-G2)"              # Stock final
ws_cmd["I2"] = '=IF(H2<Paramètres!B6,"ALERTE","OK")'

# Recopie formules pour les lignes suivantes
for r in range(3, NB_LIGNES_COMMANDE+2):
    ws_cmd[f"E{r}"] = f"=IFERROR(C{r}/D{r},0)"
    ws_cmd[f"F{r}"] = f"=H{r-1}"
    ws_cmd[f"G{r}"] = f'=IFERROR(SUMIF(Ventes!A:A,A{r},Ventes!B:B),0)'
    ws_cmd[f"H{r}"] = f"=MAX(0,F{r}+E{r}-G{r})"
    ws_cmd[f"I{r}"] = f'=IF(H{r}<Paramètres!B6,"ALERTE","OK")'

# Format conditionnel
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

rule_alert = FormulaRule(formula=['$I2="ALERTE"'], fill=red_fill)
rule_ok = FormulaRule(formula=['$I2="OK"'], fill=green_fill)

ws_cmd.conditional_formatting.add(f"I2:I{NB_LIGNES_COMMANDE+1}", rule_alert)
ws_cmd.conditional_formatting.add(f"I2:I{NB_LIGNES_COMMANDE+1}", rule_ok)

# Graphique Stock final
chart = LineChart()
chart.title = "Niveau de cuve (Stock final)"
data_ref = Reference(ws_cmd, min_col=8, min_row=1, max_row=NB_LIGNES_COMMANDE+1)
chart.add_data(data_ref, titles_from_data=True)
dates_ref = Reference(ws_cmd, min_col=1, min_row=2, max_row=NB_LIGNES_COMMANDE+1)
chart.set_categories(dates_ref)
ws_cmd.add_chart(chart, "K2")

# -------------------------
# Feuille Ventes
# -------------------------
ws_sales = wb.create_sheet("Ventes")
ws_sales["A1"] = "Date"
ws_sales["B1"] = "Quantité vendue (L)"
ws_sales["C1"] = "Montant encaissé"

for col in ["A","B","C"]:
    ws_sales.column_dimensions[col].width = 20

# -------------------------
# Sauvegarde
# -------------------------
wb.save(OUTPUT_FILE)
print(f"Fichier généré: {OUTPUT_FILE}")
